import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

product_inv_under10 = {}
product_per_supplier = {}
total_value_supplier_prod = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    price = product_list.cell(product_row, 3).value
    inventory = product_list.cell(product_row, 2).value
    product = product_list.cell(product_row, 1).value
    product_inv_value = product_list.cell(product_row, 5)
    if supplier_name in product_per_supplier:
        current_amount = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_amount + 1

    else:
        product_per_supplier[supplier_name] = 1

    if supplier_name in total_value_supplier_prod:
        current_value = total_value_supplier_prod.get(supplier_name)
        total_value_supplier_prod[supplier_name] = current_value + inventory * price

    else:
        total_value_supplier_prod[supplier_name] = inventory * price

    if inventory < 10:
        product_inv_under10[product] = inventory

    product_inv_value.value = inventory * price

print(product_per_supplier)
print(product_inv_under10)
print(total_value_supplier_prod)

inv_file.save("inventory_with_product_value.xlsx")